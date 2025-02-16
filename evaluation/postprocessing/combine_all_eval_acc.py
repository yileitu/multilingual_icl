import os
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(__file__)
parent_dir = os.path.dirname(script_dir)
grandparent_dir = os.path.dirname(parent_dir)
sys.path.insert(0, grandparent_dir)

from util.func import get_low_high_resource_langs

ROW_COLORS = [
	'E6E6FA',  # Lavender
	'FFE4E1',  # Misty Rose
	'F0E68C',  # Khaki
	'E0FFFF',  # Light Cyan
	'F0FFF0',  # Honeydew
	'FFF0F5',  # Lavender Blush
	'F5F5DC',  # Beige
	'F0F8FF',  # Alice Blue
	'F8F8FF',  # Ghost White
	'F0FFFF',  # Azure
	'FFF5EE',  # Seashell
	]


def extract_icl_language(mode_name):
	"""Extract the language part from the ICL mode name"""
	# Assuming format is always 'icl-{language}_cot-english'
	try:
		res = mode_name.split('-')[1].split('_')[0]
		if "google-translate-test-questions" in mode_name:
			res += "_google-translate-test-questions"
		if "google-translate-demonstrations" in mode_name:
			res += "_google-translate-demonstrations"
		return res
	except:
		return str(mode_name)


def extract_rand_sent_mode(mode_name):
	mode_name = str(mode_name)
	try:
		icl_lang = mode_name.split('icl-')[-1].split('_')[0]
		rand_sent_lang = mode_name.split('rand-sent-')[-1].split('-')[0]
		return f"{icl_lang}-IrrSent-{rand_sent_lang}"
	except:
		return mode_name


def apply_excel_formatting(filename):
	wb = load_workbook(filename)
	ws = wb.active

	# Format header row
	header_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
	for cell in ws[1]:
		cell.fill = header_fill
		cell.font = Font(bold=True)

	# First pass: identify group boundaries and sizes
	current_model = None
	group_boundaries = []  # Store start and end rows of each group
	start_row = 2  # Start after header

	for row in range(2, ws.max_row + 1):
		model = ws.cell(row=row, column=1).value
		if current_model is not None and model != current_model:
			group_boundaries.append((start_row, row - 1))
			start_row = row
		current_model = model

	# Add the last group
	if start_row <= ws.max_row:
		group_boundaries.append((start_row, ws.max_row))

	# Second pass: apply colors consistently across groups
	row_to_insert = []  # Track where to insert empty rows

	# Apply colors to each group
	for group_idx, (start, end) in enumerate(group_boundaries):
		# Add empty row before each group (except the first)
		if group_idx > 0:
			row_to_insert.append(start)

		# Apply colors within the group
		relative_row = 0  # Position within the group
		for row in range(start, end + 1):
			# Use relative_row to determine color, ensuring consistency across groups
			color = ROW_COLORS[relative_row % len(ROW_COLORS)]
			fill = PatternFill(start_color=color, end_color=color, fill_type='solid')

			# Apply the color to all cells in the row
			for col in range(1, ws.max_column + 1):
				ws.cell(row=row, column=col).fill = fill

			relative_row += 1

	# Insert empty rows
	for row in reversed(row_to_insert):
		ws.insert_rows(row)

	# Auto-adjust column widths
	for col in range(1, ws.max_column + 1):
		column_letter = get_column_letter(col)
		max_length = 0
		for cell in ws[column_letter]:
			try:
				if len(str(cell.value)) > max_length:
					max_length = len(str(cell.value))
			except:
				pass
		adjusted_width = (max_length + 2)
		ws.column_dimensions[column_letter].width = adjusted_width

	wb.save(filename)


def process_excel_files(root_dir, dataset_name: str, save_fpath: str, target_models=None, target_modes=None,
                        is_rand_sent: bool = False) -> pd.DataFrame:
	# Store all accuracy data with their model and mode identifiers
	all_data = []
	languages = None

	# Walk through specified models
	for model_name in target_models:
		model_dir = Path(root_dir) / model_name
		if not model_dir.exists() or not model_dir.is_dir():
			print(f"Warning: Model directory {model_name} not found")
			continue

		# Look for specified modes
		if is_rand_sent:
			modes = [d for d in model_dir.iterdir() if d.is_dir() and "rand-sent" in d.name]
			modes = sorted(modes, key=lambda x: extract_rand_sent_mode(x.name))
		else:
			modes = target_modes

		for mode_path in modes:
			mode_dir = model_dir / mode_path
			if not mode_dir.exists() or not mode_dir.is_dir():
				print(f"Warning: Mode directory {mode_path} not found in {model_name}")
				continue

			excel_file = mode_dir / f"{dataset_name.lower()}_evaluation_metrics.xlsx"

			if excel_file.exists():
				try:
					# Read the Excel file
					df = pd.read_excel(excel_file)

					# Store languages if not already stored
					if languages is None:
						languages = df['Language'].tolist()

					# Get accuracy values
					accuracies = df['Accuracy'].tolist()

					# Extract ICL language from mode name
					if is_rand_sent:
						mode_col_name = extract_rand_sent_mode(mode_path)
					else:
						mode_col_name = extract_icl_language(mode_path)

					# Create row data with separate model and mode
					row_data = {
						'Model': model_name,
						'Mode' : mode_col_name
						}

					# Add accuracy values for each language
					for lang, acc in zip(languages, accuracies):
						row_data[lang] = acc

					# Store the data
					all_data.append(row_data)
					print(f"Processed: {model_name} - {mode_path}")

				except Exception as e:
					print(f"Error processing {excel_file}: {str(e)}")
			else:
				print(f"Warning: Excel file not found in {mode_dir}")

	# Create the new dataframe
	if languages and all_data:
		# Create the final dataframe
		result_df = pd.DataFrame(all_data)

		calc_avg(languages, result_df)

		# Reorder columns to put Model and Mode first, then individual accuracies, then averages
		cols = ['Model', 'Mode'] + languages + ['Avg_Low', 'Avg_High', 'Avg_All']
		result_df = result_df[cols]

		# Save to Excel
		if is_rand_sent:
			save_fpath = save_fpath.replace("summary", "rand-sent_summary")
		output_file = Path(root_dir) / save_fpath
		result_df.to_excel(output_file, index=False)
		print(f"\nCombined data saved to {output_file}")

		apply_excel_formatting(output_file)

		# Display the first few rows
		print("\nFirst few rows of the combined data:")
		print(result_df.head())

		return result_df

	return None


def calc_avg(languages, result_df):
	low_langs, high_langs = get_low_high_resource_langs(languages)
	# Calculate averages for different language groups
	# Low-resource languages average
	result_df['Avg_Low'] = result_df[low_langs].mean(axis=1).round(4)
	# High-resource languages average
	result_df['Avg_High'] = result_df[high_langs].mean(axis=1).round(4)
	# Overall average
	result_df['Avg_All'] = result_df[languages].mean(axis=1).round(4)


# Usage example
if __name__ == "__main__":
	DATASET = "XCOPA"
	NATIVE: bool = False
	IS_RAND_SENT: bool = False
	root_directory = f"/h/yileitu/multilingual_exemplar/evaluation/{DATASET}/{DATASET.lower()}_eval"  # Replace with your actual path

	# Configuration lists
	MODELS = [
		'llama3-8b-instruct',
		'llama3.1-8b-instruct',
		'qwen2-7b-instruct',
		'qwen2.5-7b-instruct',
		'Mistral-Nemo-Instruct-2407',
		'aya-expanse-8b'
		]

	if DATASET == "MGSM":
		if NATIVE:
			MODES = [
				'icl-english_cot-native',
				'icl-french_cot-native',
				'icl-chinese_cot-native',
				'icl-japanese_cot-native',
				'icl-multilingual_cot-native',
				'icl-native_cot-native',
				]
			excel_fname: str = f"{DATASET.lower()}_cot-native_summary.xlsx"
		else:
			MODES = [
				'icl-english_cot-english',
				'icl-english_cot-english_google-translate-test-questions',
				'icl-english_cot-english_google-translate-demonstrations',
				'icl-french_cot-english',
				'icl-chinese_cot-english',
				'icl-japanese_cot-english',
				'icl-multilingual_cot-english',
				'icl-native_cot-english',
				]
			excel_fname: str = f"{DATASET.lower()}_cot-english_summary.xlsx"
	elif DATASET == "XCOPA":
		MODES = [
			'icl-english_cot-direct_all',
			'icl-english_cot-direct_all_google-translate-test-questions',
			'icl-english_cot-direct_all_google-translate-demonstrations',
			'icl-italian_cot-direct_all',
			'icl-chinese_cot-direct_all',
			'icl-multilingual_cot-direct_all',
			'icl-native_cot-direct_all',
			]
		excel_fname: str = f"{DATASET.lower()}_cot-direct_summary.xlsx"
	elif DATASET == "XLWIC":
		MODES = [
			'icl-english_cot-direct_all',
			'icl-french_cot-direct_all',
			'icl-chinese_cot-direct_all',
			'icl-japanese_cot-direct_all',
			'icl-multilingual_cot-direct_all',
			'icl-native_cot-direct_all',
			]
		excel_fname: str = f"{DATASET.lower()}_cot-direct_summary.xlsx"

	result = process_excel_files(
		root_directory,
		dataset_name=DATASET,
		save_fpath=excel_fname,
		target_models=MODELS,
		target_modes=MODES,
		is_rand_sent=IS_RAND_SENT
		)
